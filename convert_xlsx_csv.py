# import packages
import os
import pandas as pd
import numpy as np
from datetime import datetime
import time
import logging
from openpyxl import load_workbook

# setup debug logging
logging.basicConfig(level=logging.WARN)
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

# set file pathway variables an expand to HOME
path = str(os.getcwd()+'/')
print(path)
tweet_list = "Tweet_List.xlsx"
cand_tweets = "Presidential_Tweets.xlsx"
pac_tweets = "PAC_Tweets.xlsx"
path = os.path.expanduser(path)

# sheetnames
cand_sheet = 'candidate'
pac_sheet = 'pac'

def load_sheets(path):
  sheet_book = load_workbook(path)
  sheet_writer = pd.ExcelWriter(path, engine='openpyxl')
  sheet_writer.book = sheet_book
  sheet_writer.sheets = dict((ws.title, ws) for ws in sheet_book.worksheets)
  logger.info("Downloaded %s" % path)
  return sheet_writer

def convert_xlsx_csv (tweet_sheet, sheetname, tweet_list):
  # start timer
  start = time.time()
  logger.info("Start...")
  # dp_client = authenticate_dropbox()

  # load and prepare list of twitter accounts
  list_writer = load_sheets(tweet_list)
  list_df = pd.read_excel(tweet_list, sheetname=sheetname)
  list_df = list_df.dropna(thresh=4)

  #merged_corpus = pd.DataFrame(columns=['id', 'created_at', 'text', 'hashtag#', 'at@', 'link', 'retweets', 'favorites', 'full URL'])
  merged_df = pd.DataFrame()

  initial_loop = True

  # loop through the list of Cand/PACs and updates each tweet sheet appropriately
  for index, row in list_df.iterrows():
    name, since_id, count = row[1], row[2],row[3]

    if(name == 'POTUS'):
      continue

    if (initial_loop):
      merged_df = pd.read_excel(tweet_sheet, sheetname=name)
      merged_df['Name'] = name
      logger.info("Retrived data from spreadsheet for %s" % name)
      initial_loop = False

    else:
      # read current cand tweet sheet
      curr_df = pd.read_excel(tweet_sheet, sheetname=name)
      curr_df['Name'] = name
      #print (curr_df)
      logger.info("Retrived data from spreadsheet for %s" % name)

      merged_df = merged_df.append(curr_df)

  # write the updated list and save the changes to the excel sheets
  merged_df.to_csv('merged_corpus.csv', encoding='utf-8')

  logger.info("done")

convert_xlsx_csv(path + cand_tweets, cand_sheet, path + tweet_list)
